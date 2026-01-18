"""
Excel Dashboard Template - Amplifier Stories Style

Creates a professional Excel dashboard with:
- Black header styling with blue accents
- Clean data presentation
- Visual charts and metrics
- Professional formatting matching Keynote aesthetic

Usage:
    python dashboard-template.py output.xlsx "Dashboard Title" data_dict
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Amplifier Keynote Color Palette (Excel RGB format)
COLORS = {
    "black": "000000",
    "white": "FFFFFF",
    "blue": "0A84FF",
    "purple": "5E5CE6",
    "green": "30D158",
    "orange": "FF9F0A",
    "gray_dark": "1F1F1F",
    "gray_light": "F3F2F1",
    "code_green": "98D4A0",
}


def create_dashboard(filename, title, metrics_data=None):
    """
    Create a dashboard with professional styling.

    Args:
        filename: Output .xlsx filename
        title: Dashboard title
        metrics_data: Dict of metric_name: value pairs
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # === HEADER SECTION ===
    # Title row
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=24, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill(
        start_color=COLORS["black"], end_color=COLORS["black"], fill_type="solid"
    )
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle/date row
    ws["A2"] = "Amplifier Ecosystem Metrics"
    ws["A2"].font = Font(name="Arial", size=12, color=COLORS["blue"])
    ws["A2"].fill = PatternFill(
        start_color=COLORS["black"], end_color=COLORS["black"], fill_type="solid"
    )
    ws.row_dimensions[2].height = 25

    # Merge header cells
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    # === METRICS SECTION ===
    current_row = 4

    if metrics_data:
        # Section header
        ws[f"A{current_row}"] = "KEY METRICS"
        ws[f"A{current_row}"].font = Font(
            name="Arial", size=11, bold=True, color=COLORS["blue"]
        )
        ws[f"A{current_row}"].fill = PatternFill(
            start_color=COLORS["gray_light"], fill_type="solid"
        )
        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Column headers
        ws[f"A{current_row}"] = "Metric"
        ws[f"B{current_row}"] = "Value"
        ws[f"C{current_row}"] = "Change"
        ws[f"D{current_row}"] = "Target"
        ws[f"E{current_row}"] = "Status"

        for col in ["A", "B", "C", "D", "E"]:
            cell = ws[f"{col}{current_row}"]
            cell.font = Font(name="Arial", size=10, bold=True, color=COLORS["black"])
            cell.fill = PatternFill(start_color=COLORS["gray_light"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color=COLORS["black"]))

        current_row += 1

        # Sample data rows (replace with actual data)
        for metric_name, value in (metrics_data or {}).items():
            ws[f"A{current_row}"] = metric_name
            ws[f"B{current_row}"] = value
            ws[f"C{current_row}"] = (
                "=B{current_row}-B{current_row+1}"  # Example formula
            )
            ws[f"D{current_row}"] = value * 1.2  # Example target
            ws[f"E{current_row}"] = "✓"

            # Formatting
            ws[f"A{current_row}"].font = Font(
                name="Arial", size=10, color=COLORS["black"]
            )
            ws[f"B{current_row}"].font = Font(
                name="Arial", size=10, bold=True, color=COLORS["blue"]
            )
            ws[f"B{current_row}"].number_format = "#,##0"
            ws[f"E{current_row}"].font = Font(
                name="Arial", size=12, color=COLORS["green"]
            )
            ws[f"E{current_row}"].alignment = Alignment(horizontal="center")

            current_row += 1

    # === COLUMN WIDTHS ===
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 20

    # === INSTRUCTIONS SHEET ===
    instructions = wb.create_sheet("Instructions")
    instructions["A1"] = "Dashboard Instructions"
    instructions["A1"].font = Font(
        name="Arial", size=16, bold=True, color=COLORS["blue"]
    )

    instructions["A3"] = "Color Coding:"
    instructions["A4"] = "• Blue text = Formulas and calculations"
    instructions["A5"] = "• Green checkmarks = On target"
    instructions["A6"] = "• Orange warnings = Needs attention"
    instructions["A7"] = ""
    instructions["A8"] = "This dashboard uses the Amplifier Keynote style:"
    instructions["A9"] = "• Clean, professional design"
    instructions["A10"] = "• Blue accents for emphasis"
    instructions["A11"] = "• Easy-to-read metrics"

    for row in range(3, 12):
        instructions[f"A{row}"].font = Font(name="Arial", size=10)

    instructions.column_dimensions["A"].width = 50

    wb.save(filename)
    return filename


if __name__ == "__main__":
    # Example usage
    sample_metrics = {
        "Active Users": 150,
        "Features Shipped": 45,
        "Repos Created": 12,
        "Code Generated (lines)": 125000,
    }

    create_dashboard(
        "dashboard-example.xlsx", "Amplifier Adoption Dashboard", sample_metrics
    )
    print("Dashboard template created: dashboard-example.xlsx")

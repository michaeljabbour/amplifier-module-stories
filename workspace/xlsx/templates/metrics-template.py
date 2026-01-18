"""
Excel Metrics Template - Amplifier Stories Style

Creates a clean metrics table with:
- Blue accent headers
- Formula-driven calculations
- Professional number formatting
- Trend indicators

Usage:
    from metrics_template import create_metrics_sheet
    create_metrics_sheet(workbook, sheet_name, metrics_data)
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore


COLORS = {
    "blue": "0A84FF",
    "white": "FFFFFF",
    "black": "000000",
    "gray_light": "F3F2F1",
    "green": "30D158",
    "orange": "FF9F0A",
    "red": "FF453A",
}


def create_metrics_sheet(wb, sheet_name, metrics):  # type: ignore
    """
    Create a metrics tracking sheet.

    Args:
        wb: Workbook object
        sheet_name: Name for the sheet
        metrics: List of dicts with 'name', 'current', 'previous', 'target'
    """
    ws = wb.create_sheet(sheet_name)

    # Header
    ws["A1"] = sheet_name.upper()
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=COLORS["blue"])
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = ["Metric", "Current", "Previous", "Target", "Status"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(name="Arial", size=11, bold=True, color=COLORS["black"])
        cell.fill = PatternFill(start_color=COLORS["gray_light"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=COLORS["blue"]))

    ws.row_dimensions[3].height = 25

    # Data rows
    for idx, metric in enumerate(metrics, start=4):
        ws[f"A{idx}"] = metric.get("name", "")
        ws[f"B{idx}"] = metric.get("current", 0)
        ws[f"C{idx}"] = metric.get("previous", 0)
        ws[f"D{idx}"] = metric.get("target", 0)

        # Calculate status with formula
        ws[f"E{idx}"] = f'=IF(B{idx}>=D{idx},"✓",IF(B{idx}>=C{idx},"→","↓"))'

        # Formatting
        ws[f"A{idx}"].font = Font(name="Arial", size=10)
        ws[f"B{idx}"].font = Font(
            name="Arial", size=10, bold=True, color=COLORS["blue"]
        )
        ws[f"B{idx}"].number_format = "#,##0"
        ws[f"C{idx}"].number_format = "#,##0"
        ws[f"D{idx}"].number_format = "#,##0"
        ws[f"E{idx}"].alignment = Alignment(horizontal="center")
        ws[f"E{idx}"].font = Font(name="Arial", size=12)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 10

    return ws

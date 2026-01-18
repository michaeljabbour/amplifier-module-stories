"""
Excel Comparison Template - Amplifier Stories Style

Creates a before/after comparison table with:
- Side-by-side metrics
- Visual improvement indicators
- Clean formatting
- Formula-driven calculations

Usage:
    from comparison_template import create_comparison_sheet
    create_comparison_sheet(workbook, "Feature Impact", before_data, after_data)
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore


COLORS = {
    "blue": "0A84FF",
    "green": "30D158",
    "orange": "FF9F0A",
    "gray_light": "F3F2F1",
    "black": "000000",
}


def create_comparison_sheet(wb, title, before_metrics, after_metrics):  # type: ignore
    """
    Create before/after comparison sheet.

    Args:
        wb: Workbook object
        title: Sheet title
        before_metrics: Dict of metric_name: value (before)
        after_metrics: Dict of metric_name: value (after)
    """
    ws = wb.create_sheet(title)

    # Header
    ws["A1"] = title.upper()
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=COLORS["blue"])
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30

    # Column headers
    ws["A3"] = "Metric"
    ws["B3"] = "Before"
    ws["C3"] = "After"
    ws["D3"] = "Improvement"

    for col in ["A", "B", "C", "D"]:
        cell = ws[f"{col}3"]
        cell.font = Font(name="Arial", size=11, bold=True, color=COLORS["black"])
        cell.fill = PatternFill(start_color=COLORS["gray_light"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=COLORS["blue"]))

    ws.row_dimensions[3].height = 25

    # Data rows
    row = 4
    for metric_name in before_metrics.keys():
        ws[f"A{row}"] = metric_name
        ws[f"B{row}"] = before_metrics[metric_name]
        ws[f"C{row}"] = after_metrics.get(metric_name, 0)
        ws[f"D{row}"] = f'=IF(C{row}>B{row},(C{row}-B{row})/B{row}*100&"%","-")'

        # Formatting
        ws[f"A{row}"].font = Font(name="Arial", size=10)
        ws[f"B{row}"].number_format = "#,##0"
        ws[f"C{row}"].number_format = "#,##0"
        ws[f"C{row}"].font = Font(
            name="Arial", size=10, bold=True, color=COLORS["green"]
        )
        ws[f"D{row}"].font = Font(name="Arial", size=10, color=COLORS["blue"])
        ws[f"D{row}"].alignment = Alignment(horizontal="right")

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    return ws
